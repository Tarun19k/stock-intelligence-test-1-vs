#!/usr/bin/env python3
"""
ingest_holdings_from_pnl.py — parse a real broker P&L export (Groww format,
"Trade Level" + "Scrip Level" sheets) into the `holdings` table.

Source file (real, personal financial data): Stocks_PnL_Report.xlsx.
Real writes are gated: this script is DRY-RUN by default. Pass --live-write
--i-understand-this-writes-real-financial-data to actually insert.

Data shape, confirmed live by direct read of the real file:
- "Scrip Level" sheet, "Unrealised (Holdings as on <date>)" section: one row per
  distinct stock, ALREADY aggregated (qty, weighted avg buy price) by the broker.
  This is the right granularity for the current single-row-per-ticker `holdings`
  schema.
- "Trade Level" sheet, "Unrealised trades" section: one row per BUY LOT (same
  stock can appear multiple times with different buy dates/prices) for currently
  held positions. Used here only to derive `acquired_at` (earliest lot date per
  stock) since the aggregated Scrip Level sheet has no date column.

KNOWN LIMITATION, flagged not hidden: using earliest-lot-date as `acquired_at`
and a single blended avg_cost loses true per-lot FIFO detail. This is fine for
Phase C's portfolio-value/concentration diagnostics, but Prereq 7's tax logic
(grandfathering, holding-period-per-lot for STCG/LTCG classification) needs
real per-lot detail this schema does not currently capture. A `holdings_lots`
table is a real, separate future need if precise tax-lot accounting is required
before this is used to compute the Harvest engine's tax-loss/gain candidates.
"""

import argparse
import sys
from datetime import datetime, timezone

import openpyxl
import requests

import re

from ingest_fundamentals import _fetch_scrip_master, _validate_scrip_master

PNL_PATH = "/Users/home/Tarun's team Dropbox/Tarun Kochhar/Mac (2)/Downloads/Stocks_PnL_Report.xlsx"


def _normalize_name(name: str) -> str:
    """Strip legal-entity suffixes and punctuation for fuzzy name matching.
    No hardcoded company list -- generic normalization only, so this scales to
    any future stock without a per-name entry."""
    name = name.upper()
    name = re.sub(r"\b(LTD|LIMITED|LIL|LLP|PVT|PRIVATE|CORP|CORPORATION|CO|COMPANY)\.?\b", "", name)
    name = re.sub(r"[^A-Z0-9]", "", name)
    return name


def resolve_tickers():
    """Live BSE scrip-master ISIN -> real trading symbol (scrip_id) resolution,
    plus a normalized-name index for fallback matching when the ISIN in a source
    file doesn't match the current scrip master (real case found 2026-07-30: an
    ISIN reissue, e.g. after a reconstruction scheme, leaves an old export with a
    stale ISIN even though the company/ticker itself is unchanged and resolvable
    by name). No hardcoded per-company mapping -- this generalizes to any stock,
    not just the one case found so far."""
    rows = _validate_scrip_master(_fetch_scrip_master())
    isin_index = {r["ISIN_NUMBER"]: r["scrip_id"] for r in rows if r.get("ISIN_NUMBER") and r.get("scrip_id")}
    name_index = {}
    for r in rows:
        if not r.get("scrip_id"):
            continue
        for field in ("Scrip_Name", "Issuer_Name"):
            if r.get(field):
                name_index[_normalize_name(r[field])] = r["scrip_id"]
    return isin_index, name_index


def resolve_one_ticker(name: str, isin: str, isin_index: dict, name_index: dict):
    """Returns (ticker, verified, method). Tries ISIN first (exact, strongest
    signal), falls back to normalized-name match (handles ISIN reissues), falls
    back to an unverified name-derived placeholder (flagged, never silent)."""
    if isin in isin_index:
        return isin_index[isin], True, "isin"
    normalized = _normalize_name(name)
    if normalized in name_index:
        return name_index[normalized], True, "name_fallback"
    return make_fallback_ticker(name), False, "unresolved"


def parse_scrip_level_unrealised(ws):
    """Returns {stock_name: {isin, qty, avg_cost}} from the Scrip Level sheet's
    Unrealised section."""
    holdings = {}
    in_section = False
    header_seen = False
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0] and "Unrealised" in str(row[0]):
            in_section = True
            continue
        if not in_section:
            continue
        if row[0] == "Stock name":
            header_seen = True
            continue
        if header_seen and row[0]:
            name, isin, qty, avg_buy = row[0], row[1], row[2], row[3]
            if qty and qty > 0:
                holdings[name] = {"isin": isin, "qty": qty, "avg_cost": avg_buy}
    return holdings


def parse_trade_level_earliest_dates(ws):
    """Returns {stock_name: earliest_buy_date (date obj)} from the Trade Level
    sheet's Unrealised trades section."""
    dates = {}
    in_section = False
    header_seen = False
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0] and "Unrealised trades" in str(row[0]):
            in_section = True
            continue
        if not in_section:
            continue
        if row[0] == "Stock name":
            header_seen = True
            continue
        if header_seen and row[0]:
            name, buy_date_str = row[0], row[3]
            if buy_date_str:
                try:
                    d = datetime.strptime(str(buy_date_str), "%d-%m-%Y").date()
                except ValueError:
                    continue
                if name not in dates or d < dates[name]:
                    dates[name] = d
    return dates


def make_fallback_ticker(stock_name: str) -> str:
    """Fallback only, used when live BSE resolution fails (ETF/MF units, REITs,
    delisted/suspended equity -- not in BSE's Equity-segment scrip master).
    NOT a verified NSE/BSE ticker symbol -- flagged in the row itself."""
    return stock_name.strip().upper().replace(" ", "_").replace(".", "").replace("-", "_")[:30]


def build_rows(pnl_path: str):
    wb = openpyxl.load_workbook(pnl_path, data_only=True)
    scrip_holdings = parse_scrip_level_unrealised(wb["Scrip Level"])
    earliest_dates = parse_trade_level_earliest_dates(wb["Trade Level"])
    isin_index, name_index = resolve_tickers()

    rows = []
    missing_dates = []
    unresolved_tickers = []
    for name, info in scrip_holdings.items():
        acquired_at = earliest_dates.get(name)
        if not acquired_at:
            missing_dates.append(name)
            continue

        isin = info["isin"]
        ticker, verified, method = resolve_one_ticker(name, isin, isin_index, name_index)
        if not verified:
            unresolved_tickers.append((name, isin))

        rows.append({
            "ticker": ticker,
            "qty": info["qty"],
            "avg_cost": round(info["avg_cost"], 2),
            "acquired_at": acquired_at.isoformat(),
            "source": "groww_pnl_export",
            "ingested_at": datetime.now(timezone.utc).isoformat(),
            "_real_name": name,
            "_isin": isin,
            "_ticker_verified": verified,
            "_resolution_method": method,
        })
    return rows, missing_dates, unresolved_tickers


def build_lot_rows(pnl_path: str, isin_by_name: dict):
    """Real per-lot rows for holdings_lots, from Trade Level's Unrealised trades."""
    wb = openpyxl.load_workbook(pnl_path, data_only=True)
    ws = wb["Trade Level"]
    isin_index, name_index = resolve_tickers()

    lots = []
    in_section = False
    header_seen = False
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0] and "Unrealised trades" in str(row[0]):
            in_section = True
            continue
        if not in_section:
            continue
        if row[0] == "Stock name":
            header_seen = True
            continue
        if header_seen and row[0]:
            name, isin, qty, buy_date_str, buy_price = row[0], row[1], row[2], row[3], row[4]
            if not qty or qty <= 0:
                continue
            try:
                buy_date = datetime.strptime(str(buy_date_str), "%d-%m-%Y").date()
            except (ValueError, TypeError):
                continue
            ticker, _, _ = resolve_one_ticker(name, isin, isin_index, name_index)
            lots.append({
                "ticker": ticker,
                "isin": isin,
                "real_name": name,
                "qty": qty,
                "buy_price": round(buy_price, 4) if buy_price else 0,
                "buy_date": buy_date.isoformat(),
                "source": "groww_pnl_export",
                "ingested_at": datetime.now(timezone.utc).isoformat(),
            })
    return lots


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--live-write", action="store_true")
    parser.add_argument("--i-understand-this-writes-real-financial-data", action="store_true")
    args = parser.parse_args()

    rows, missing, unresolved = build_rows(PNL_PATH)
    isin_by_name = {r["_real_name"]: r["_isin"] for r in rows}
    lots = build_lot_rows(PNL_PATH, isin_by_name)

    verified = sum(1 for r in rows if r["_ticker_verified"])
    by_isin = sum(1 for r in rows if r["_resolution_method"] == "isin")
    by_name = sum(1 for r in rows if r["_resolution_method"] == "name_fallback")
    print(f"Parsed {len(rows)} current holdings ({verified}/{len(rows)} tickers verified: "
          f"{by_isin} by ISIN, {by_name} by name-fallback) and {len(lots)} real per-lot records.")
    if missing:
        print(f"WARNING: {len(missing)} stocks had no matching buy-date, skipped: {missing}")
    if unresolved:
        print(f"UNRESOLVED tickers ({len(unresolved)}, using name-derived fallback, "
              f"NOT verified against BSE — mostly ETF/MF units or delisted/suspended equity):")
        for n, i in unresolved:
            print(f"  {n} ({i})")

    total_value = sum(r["qty"] * r["avg_cost"] for r in rows)
    print(f"\nTotal cost-basis value across {len(rows)} holdings: Rs.{total_value:,.2f}")

    if not args.live_write:
        print("\nDRY RUN ONLY -- nothing written. Pass --live-write "
              "--i-understand-this-writes-real-financial-data to insert for real.")
        return

    if not args.i_understand_this_writes_real_financial_data:
        print("\n--live-write requires --i-understand-this-writes-real-financial-data. Aborting.")
        sys.exit(1)

    url = key = None
    with open("../../.env") as f:
        for line in f:
            if line.startswith("SUPABASE_URL"):
                url = line.strip().split("=", 1)[1]
            if line.startswith("SUPABASE_SERVICE_KEY"):
                key = line.strip().split("=", 1)[1]

    headers = {"apikey": key, "Authorization": f"Bearer {key}",
               "Content-Type": "application/json", "Prefer": "return=representation"}

    holdings_payload = [{k: v for k, v in r.items() if not k.startswith("_")} for r in rows]
    resp1 = requests.post(f"{url}/rest/v1/holdings", headers=headers, json=holdings_payload)
    print(f"\nLIVE WRITE (holdings): HTTP {resp1.status_code}")
    print(resp1.text[:300])

    resp2 = requests.post(f"{url}/rest/v1/holdings_lots", headers=headers, json=lots)
    print(f"\nLIVE WRITE (holdings_lots): HTTP {resp2.status_code}")
    print(resp2.text[:300])


if __name__ == "__main__":
    main()
